from openpyxl import load_workbook
print('-' * 20)
print('''Orientações para uso:
*Este programa copia automaticamente dados de arquivos no formato .txt para arquivos no formato .xlsx;
*Você precisa criar arquivos com 'nomes base' seguidos de um número sequencial, ou seja, todos os arquivos de mesmo formato devem ter o mesmo\nnome e variar apenas no número;
Por exemplo: suponha que você tenha vários arquivos xlsx para preencher com números de telefone.
Em seu diretório, crie um arquivo com uma planilha modelo, e o copie até dar o número de arquivos que precisa.
Nomeie os arquivos do seguinte modo: telefones01.xlsx, telefones02.xlsx e assim por diante.
Faça o mesmo com os arquivos .txt.
É importante que nos arquivos txt os dados estejam 'juntos', isto é, que estejam um embaixo do outro sem linhas em branco. 
Este programa irá pegar os dados linha por linha, por isso os organize com antecedência nos arquivos de textos de modo que o programa só precise pegar os dados e copiá-los nas células fornecidas.
*Com os passos acima\nconcluídos, basta informar o menor número sequencial (ponto de início) e o maior número (fim) dos arquivos xlsx. Vale lembrar que o arquivo do\nnúmero que marca o fim também será preenchido!
*Certifique-se de que o nome das planilhas dentro de todos os arquivos xlsx sejam IGUAIS.''')

print('-' * 20, 'DADOS DO DIRETÓRIO', '-' * 20)

nome_arquivo_xlsx = input('Nome base (sem o nº nem extensão) do arquivo .XLSX que deseja usar: ')
nome_arquivo_txt =  input('Nome base (sem o nº nem extensão) do arquivo .TXT que deseja usar: ')
num_minimo_xlsx = int(input(f'MENOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
num_maximo_xlsx = int(input(f'MAIOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
while num_maximo_xlsx <= num_minimo_xlsx:
    num_maximo_xlsx = int(input('O número do ponto de finalização deve ser MAIOR que o número do ponto de início!\nPor favor, insira o MAIOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
num_arquivo_txt = int(input(f'Número do arquivo {nome_arquivo_txt} a partir do qual deseja obter os dados: '))

print('-' * 58)
print('-' * 20, 'DADOS DA PLANILHA', '-' * 20)

colunas = input('Informe a(s) letra(s) [separando-as apenas por espaços] da(s) coluna(s) que será(ão) preenchida(s): ').strip().upper().split()
#print(colunas)
#print(len(colunas))
linhas = input('Informe os números [separando-os apenas por espaços] das linhas que serão preenchidas: ').strip().upper().split()
#print(linhas)
print('-' * 58)
                             
while num_maximo_xlsx >= num_minimo_xlsx:
    try:
        wb = load_workbook(f'{nome_arquivo_xlsx + str(num_minimo_xlsx)}.xlsx')
        nomePlan = wb.sheetnames[0]
        ws = wb[nomePlan]
        arquivo = open(f'{nome_arquivo_txt + str(num_arquivo_txt)}.txt', 'r', encoding='utf-8')
        arquivo.seek(0, 0)       
        col = 0
        while col <= len(colunas) - 1:
            for r in range(0, len(linhas)):
                ws[f'{colunas[col] + linhas[r]}'] = arquivo.readline()
                #print(f'{colunas[col] + linhas[r]}')
            col += 1
        arquivo.close()
        wb.save(f'{nome_arquivo_xlsx + str(num_minimo_xlsx)}.xlsx')
        num_minimo_xlsx += 1
        num_arquivo_txt += 1
        print('-' * 5, 'Dados copiados com sucesso!', '-' * 5)
    
    except:
        print("ALGO DEU ERRADO! CONFIRA SE PREENCHEU TUDO CORRETAMENTE E TENTE DE NOVO.")
print('-' * 15, 'Programa Finalizado', '-' * 15)
