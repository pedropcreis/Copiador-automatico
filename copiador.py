from openpyxl import load_workbook
print('-' * 15, 'COPIADOR AUTOMÁTICO PARA GRANDES QUANTIDADES DE DADOS' , '-' * 15)

print('-' * 20, 'DADOS DO DIRETÓRIO', '-' * 20)

nome_arquivo_xlsx = input('Nome base (sem o nº nem extensão) do arquivo .XLSX que deseja usar: ')
nome_arquivo_txt =  input('Nome base (sem o nº nem extensão) do arquivo .TXT que deseja usar: ')
num_minimo_xlsx = int(input(f'MENOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
num_maximo_xlsx = int(input(f'MAIOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
while num_maximo_xlsx <= num_minimo_xlsx:
    num_maximo_xlsx = int(input('O número do ponto de finalização deve ser MAIOR que o número do ponto de início!\nPor favor, insira o MAIOR número do arquivo {nome_arquivo_xlsx} que deseja editar: '))
num_arquivo_txt = int(input(f'Número do arquivo {nome_arquivo_txt} a partir do qual deseja obter os dados: '))

print('-' * 58)
print('-' * 20, 'DADOS DA PLANILHA', '-' * 20)

colunas = input('Informe a(s) letra(s) [separando-as apenas por espaços] da(s) coluna(s) que será(ão) preenchida(s): ').strip().upper().split()
#print(colunas)
#print(len(colunas))
linhas = input('Informe os números [separando-os apenas por espaços] das linhas que serão preenchidas: ').strip().upper().split()
#print(linhas)
print('-' * 58)
                             
while num_maximo_xlsx >= num_minimo_xlsx:
    try:
        wb = load_workbook(f'{nome_arquivo_xlsx + str(num_minimo_xlsx)}.xlsx')
        nomePlan = wb.sheetnames[0]
        ws = wb[nomePlan]
        arquivo = open(f'{nome_arquivo_txt + str(num_arquivo_txt)}.txt', 'r', encoding='utf-8')
        arquivo.seek(0, 0)       
        col = 0
        while col <= len(colunas) - 1:
            for r in range(0, len(linhas)):
                ws[f'{colunas[col] + linhas[r]}'] = arquivo.readline()
                #print(f'{colunas[col] + linhas[r]}')
            col += 1
        arquivo.close()
        wb.save(f'{nome_arquivo_xlsx + str(num_minimo_xlsx)}.xlsx')
        num_minimo_xlsx += 1
        num_arquivo_txt += 1
        print('-' * 5, 'Dados copiados com sucesso!', '-' * 5)
    
    except:
        print("ALGO DEU ERRADO! CONFIRA SE PREENCHEU TUDO CORRETAMENTE E TENTE DE NOVO.")
print('-' * 15, 'Programa Finalizado', '-' * 15)
